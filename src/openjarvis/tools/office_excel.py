"""office_excel.py — Excel (.xlsx) read/find/write for personal PC use.

Actions:
1. sheets — list sheet names
2. read — preview sheet / range as TSV
3. find — search query in cell values
4. write — set cells from JSON map, save to output_path (template safe)
5. create — new workbook with optional sheet + values

Empty path → last opened .xlsx/.xlsm from session_context when available.
Uses openpyxl (no Excel COM). .xls (legacy) is not supported.
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openjarvis.core.registry import ToolRegistry
from openjarvis.core.types import ToolResult
from openjarvis.tools._stubs import BaseTool, ToolSpec

_DEFAULT_MAX_CHARS = 50_000
_DEFAULT_MAX_ROWS = 200
_XLSX_SUFFIXES = {".xlsx", ".xlsm"}
_ACTIONS = ("sheets", "read", "find", "write", "create")
_CELL_REF_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?(\d{1,7})$")


def _fail(msg: str) -> ToolResult:
    return ToolResult(tool_name="office_excel", content=msg, success=False)


def _ok(msg: str, **metadata: Any) -> ToolResult:
    return ToolResult(
        tool_name="office_excel",
        content=msg,
        success=True,
        metadata=metadata or None,
    )


def _note(path: Path) -> None:
    try:
        from openjarvis.tools.session_context import note_opened

        note_opened(str(path.resolve()))
    except Exception:
        pass


def _default_output(stem: str) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = f"{stem}_{stamp}.xlsx"
    if Path("D:/").exists():
        return Path("D:/") / name
    docs = Path.home() / "Documents"
    docs.mkdir(parents=True, exist_ok=True)
    return docs / name


def _resolve_xlsx_path(
    raw: str, *, must_exist: bool = True
) -> tuple[Optional[Path], Optional[str]]:
    path_str = (raw or "").strip()
    if not path_str:
        try:
            from openjarvis.tools.session_context import get_last_opened

            last = get_last_opened()
        except Exception:
            last = None
        if last and Path(str(last)).suffix.lower() in _XLSX_SUFFIXES:
            path_str = str(last)
        else:
            return None, "No path provided and no last opened Excel file in session."

    path = Path(path_str)
    if path.suffix.lower() == ".xls":
        return None, "Legacy .xls not supported. Save as .xlsx and retry."
    if path.suffix.lower() not in _XLSX_SUFFIXES:
        return None, f"Not an Excel .xlsx/.xlsm file: {path_str}"

    from openjarvis.security.file_policy import is_sensitive_file

    if is_sensitive_file(path):
        return None, f"Access denied: {path_str} is a sensitive file."
    if must_exist and not path.exists():
        return None, f"File not found: {path_str}"
    return path, None


def _import_openpyxl() -> tuple[Any, Optional[str]]:
    try:
        import openpyxl
    except ImportError:
        return None, (
            "openpyxl not installed. Install with: uv sync --extra excel"
        )
    return openpyxl, None


def _parse_values(raw: Any) -> tuple[Optional[dict[str, Any]], Optional[str]]:
    if raw is None:
        return None, "values is required (JSON object cell→value, e.g. {\"A1\":\"x\"})."
    if isinstance(raw, dict):
        return dict(raw), None
    text = str(raw).strip()
    if not text:
        return None, "values is empty."
    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        return None, f"values must be JSON object: {exc}"
    if not isinstance(data, dict):
        return None, "values must be a JSON object (cell ref → value)."
    return data, None


def _cell_display(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    return str(value)


def _pick_sheet(wb: Any, sheet: Optional[str]) -> tuple[Any, Optional[str]]:
    name = (sheet or "").strip()
    if not name:
        return wb.active, None
    if name not in wb.sheetnames:
        return None, (
            f"Sheet {name!r} not found. Available: {', '.join(wb.sheetnames)}"
        )
    return wb[name], None


def _sheet_to_tsv(
    ws: Any,
    *,
    range_addr: Optional[str],
    max_rows: int,
    max_chars: int,
) -> str:
    if range_addr:
        rows_iter = ws[range_addr]
        # ws['A1:C3'] → tuple of rows; ws['A1'] → single cell
        if not isinstance(rows_iter, tuple):
            rows_iter = ((rows_iter,),)
        elif rows_iter and not isinstance(rows_iter[0], tuple):
            rows_iter = (rows_iter,)
    else:
        rows_iter = ws.iter_rows(
            min_row=1,
            max_row=max(1, min(ws.max_row or 1, max_rows)),
            max_col=max(1, ws.max_column or 1),
        )

    lines: list[str] = []
    for i, row in enumerate(rows_iter):
        if not range_addr and i >= max_rows:
            lines.append(f"... truncated after {max_rows} rows")
            break
        cells = []
        for cell in row:
            cells.append(_cell_display(getattr(cell, "value", cell)))
        lines.append("\t".join(cells))

    text = "\n".join(lines)
    if len(text) > max_chars:
        text = text[:max_chars] + "\n\n[Content truncated]"
    return text


def _validate_cell_ref(ref: str) -> Optional[str]:
    ref = ref.strip()
    if not _CELL_REF_RE.match(ref):
        return f"Invalid cell reference {ref!r} (use A1, B12, ...)."
    return None


@ToolRegistry.register("office_excel")
class OfficeExcelTool(BaseTool):
    """Personal-PC Excel tool: sheets, read, find, write, create."""

    tool_id = "office_excel"

    @property
    def spec(self) -> ToolSpec:
        return ToolSpec(
            name="office_excel",
            description=(
                "Excel .xlsx/.xlsm на этом ПК (openpyxl). "
                "Когда использовать: список листов → sheets; прочитать таблицу "
                "→ read; найти значение → find; записать ячейки → write "
                "(values JSON, лучше новый output_path); создать файл → create. "
                "Когда НЕ использовать: просто открыть в Excel → open_path; "
                "PDF → office_pdf; Word → office_word; старый .xls не поддерживается. "
                "path можно опустить, если .xlsx уже в КОНТЕКСТЕ ФАЙЛОВ."
            ),
            parameters={
                "type": "object",
                "properties": {
                    "action": {
                        "type": "string",
                        "description": "sheets | read | find | write | create",
                    },
                    "path": {
                        "type": "string",
                        "description": "Path to .xlsx/.xlsm (alias: file_path).",
                    },
                    "file_path": {
                        "type": "string",
                        "description": "Alias for path.",
                    },
                    "sheet": {
                        "type": "string",
                        "description": "Sheet name (default: active / first).",
                    },
                    "range": {
                        "type": "string",
                        "description": "A1 range for read, e.g. 'A1:D20'.",
                    },
                    "query": {
                        "type": "string",
                        "description": "Search string for action=find.",
                    },
                    "values": {
                        "type": "string",
                        "description": (
                            'JSON object of cell→value, e.g. {"A1":"Name","B2":10}.'
                        ),
                    },
                    "output_path": {
                        "type": "string",
                        "description": (
                            "Where to save for write/create. "
                            "For write defaults to new file (template not overwritten)."
                        ),
                    },
                    "max_rows": {
                        "type": "integer",
                        "description": "Max rows for read (default 200).",
                    },
                    "max_chars": {
                        "type": "integer",
                        "description": "Max characters to return.",
                    },
                    "overwrite": {
                        "type": "boolean",
                        "description": (
                            "For write: if true, allow saving back to path."
                        ),
                    },
                },
                "required": ["action"],
            },
            category="office",
            required_capabilities=["file:read"],
        )

    def execute(self, **params: Any) -> ToolResult:
        action = str(params.get("action") or "").strip().lower()
        if action not in _ACTIONS:
            return _fail(
                f"Unknown action {action!r}. Use: {', '.join(_ACTIONS)}"
            )
        if action == "create":
            return self._create(params)
        if action == "write":
            return self._write(params)
        if action == "sheets":
            return self._sheets(params)
        if action == "find":
            return self._find(params)
        return self._read(params)

    def _open_wb(self, path: Path, *, data_only: bool = False) -> tuple[Any, Optional[str]]:
        openpyxl, err = _import_openpyxl()
        if err:
            return None, err
        try:
            return openpyxl.load_workbook(str(path), data_only=data_only), None
        except Exception as exc:
            return None, f"Cannot open workbook: {exc}"

    def _sheets(self, params: Any) -> ToolResult:
        path, err = _resolve_xlsx_path(
            str(params.get("path") or params.get("file_path") or "")
        )
        if err or path is None:
            return _fail(err or "No path.")
        wb, err = self._open_wb(path)
        if err or wb is None:
            return _fail(err or "Open failed.")
        try:
            names = list(wb.sheetnames)
            active = wb.active.title if wb.active is not None else ""
        finally:
            wb.close()
        _note(path)
        lines = [f"Sheets in {path} ({len(names)}):"]
        for n in names:
            mark = " (active)" if n == active else ""
            lines.append(f"- {n}{mark}")
        return _ok(
            "\n".join(lines),
            file_path=str(path.resolve()),
            sheets=names,
            active=active,
        )

    def _read(self, params: Any) -> ToolResult:
        path, err = _resolve_xlsx_path(
            str(params.get("path") or params.get("file_path") or "")
        )
        if err or path is None:
            return _fail(err or "No path.")
        wb, err = self._open_wb(path, data_only=True)
        if err or wb is None:
            # data_only needs prior Excel calc; fall back to formulas/raw.
            wb, err = self._open_wb(path, data_only=False)
            if err or wb is None:
                return _fail(err or "Open failed.")

        max_rows = int(params.get("max_rows") or _DEFAULT_MAX_ROWS)
        max_chars = int(params.get("max_chars") or _DEFAULT_MAX_CHARS)
        range_addr = str(params.get("range") or "").strip() or None
        try:
            ws, serr = _pick_sheet(wb, params.get("sheet"))
            if serr or ws is None:
                return _fail(serr or "No sheet.")
            body = _sheet_to_tsv(
                ws,
                range_addr=range_addr,
                max_rows=max_rows,
                max_chars=max_chars,
            )
            sheet_name = ws.title
        except Exception as exc:
            return _fail(f"Read failed: {exc}")
        finally:
            wb.close()

        _note(path)
        header = f"Sheet {sheet_name!r} from {path}"
        if range_addr:
            header += f" range {range_addr}"
        text = f"{header}\n{body}" if body else f"{header}\n(empty)"
        return _ok(
            text,
            file_path=str(path.resolve()),
            sheet=sheet_name,
        )

    def _find(self, params: Any) -> ToolResult:
        query = str(params.get("query") or "").strip()
        if not query:
            return _fail("query is required for action=find.")
        path, err = _resolve_xlsx_path(
            str(params.get("path") or params.get("file_path") or "")
        )
        if err or path is None:
            return _fail(err or "No path.")
        wb, err = self._open_wb(path, data_only=True)
        if err or wb is None:
            wb, err = self._open_wb(path, data_only=False)
            if err or wb is None:
                return _fail(err or "Open failed.")

        max_chars = int(params.get("max_chars") or _DEFAULT_MAX_CHARS)
        sheet_filter = str(params.get("sheet") or "").strip()
        q = query.lower()
        hits: list[str] = []
        try:
            sheets = [sheet_filter] if sheet_filter else list(wb.sheetnames)
            for name in sheets:
                if name not in wb.sheetnames:
                    return _fail(
                        f"Sheet {name!r} not found. "
                        f"Available: {', '.join(wb.sheetnames)}"
                    )
                ws = wb[name]
                for row in ws.iter_rows(
                    min_row=1,
                    max_row=ws.max_row or 1,
                    max_col=ws.max_column or 1,
                ):
                    for cell in row:
                        val = _cell_display(cell.value)
                        if val and q in val.lower():
                            hits.append(f"{name}!{cell.coordinate}: {val}")
                            if len(hits) >= 200:
                                break
                    if len(hits) >= 200:
                        break
                if len(hits) >= 200:
                    break
        finally:
            wb.close()

        _note(path)
        if not hits:
            return _ok(
                f"No matches for {query!r} in {path}.",
                file_path=str(path.resolve()),
                match_count=0,
            )
        body = f"Matches for {query!r} ({len(hits)}):\n" + "\n".join(hits)
        if len(body) > max_chars:
            body = body[:max_chars] + "\n\n[Content truncated]"
        return _ok(
            body,
            file_path=str(path.resolve()),
            match_count=len(hits),
        )

    def _write(self, params: Any) -> ToolResult:
        path, err = _resolve_xlsx_path(
            str(params.get("path") or params.get("file_path") or "")
        )
        if err or path is None:
            return _fail(err or "No path.")
        values, verr = _parse_values(params.get("values"))
        if verr or values is None:
            return _fail(verr or "Invalid values.")
        if not values:
            return _fail("values is empty.")

        for ref in values:
            cerr = _validate_cell_ref(str(ref))
            if cerr:
                return _fail(cerr)

        overwrite = bool(params.get("overwrite"))
        out_raw = str(params.get("output_path") or "").strip()
        if out_raw:
            out_path = Path(out_raw)
        elif overwrite:
            out_path = path
        else:
            out_path = _default_output(path.stem + "_edit")

        if out_path.resolve() == path.resolve() and not overwrite and not out_raw:
            # default_output should never equal path; belt-and-suspenders
            out_path = _default_output(path.stem + "_edit")

        if (
            out_path.resolve() == path.resolve()
            and not overwrite
            and out_raw
        ):
            return _fail(
                "Refusing to overwrite source. Pass overwrite=true "
                "or a different output_path."
            )

        if out_path.suffix.lower() not in _XLSX_SUFFIXES:
            out_path = out_path.with_suffix(".xlsx")
        out_path.parent.mkdir(parents=True, exist_ok=True)

        wb, err = self._open_wb(path, data_only=False)
        if err or wb is None:
            return _fail(err or "Open failed.")

        try:
            ws, serr = _pick_sheet(wb, params.get("sheet"))
            if serr or ws is None:
                return _fail(serr or "No sheet.")
            for ref, value in values.items():
                ws[str(ref).strip()] = value
            wb.save(str(out_path))
            sheet_name = ws.title
        except Exception as exc:
            return _fail(f"Write failed: {exc}")
        finally:
            wb.close()

        _note(out_path)
        return _ok(
            f"Wrote {len(values)} cell(s) on sheet {sheet_name!r} → "
            f"{out_path.resolve()}",
            file_path=str(out_path.resolve()),
            source=str(path.resolve()),
            sheet=sheet_name,
            cells_set=len(values),
        )

    def _create(self, params: Any) -> ToolResult:
        openpyxl, err = _import_openpyxl()
        if err:
            return _fail(err)

        out_raw = str(
            params.get("output_path")
            or params.get("path")
            or params.get("file_path")
            or ""
        ).strip()
        out_path = Path(out_raw) if out_raw else _default_output("workbook")
        if out_path.suffix.lower() not in _XLSX_SUFFIXES:
            out_path = out_path.with_suffix(".xlsx")
        if out_path.exists() and not bool(params.get("overwrite")):
            return _fail(
                f"File exists: {out_path}. Pass overwrite=true or another path."
            )
        out_path.parent.mkdir(parents=True, exist_ok=True)

        values: dict[str, Any] = {}
        if params.get("values") is not None:
            values, verr = _parse_values(params.get("values"))
            if verr or values is None:
                return _fail(verr or "Invalid values.")
            for ref in values:
                cerr = _validate_cell_ref(str(ref))
                if cerr:
                    return _fail(cerr)

        sheet_name = str(params.get("sheet") or "Sheet1").strip() or "Sheet1"
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name[:31]
            for ref, value in values.items():
                ws[str(ref).strip()] = value
            wb.save(str(out_path))
            wb.close()
        except Exception as exc:
            return _fail(f"Create failed: {exc}")

        _note(out_path)
        return _ok(
            f"Created workbook: {out_path.resolve()} "
            f"(sheet {sheet_name!r}, {len(values)} cell(s))",
            file_path=str(out_path.resolve()),
            sheet=sheet_name,
            cells_set=len(values),
        )


__all__ = ["OfficeExcelTool"]
